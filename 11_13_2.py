import openpyxl, sys
script = sys.argv[3]
m = int(sys.argv[1])
n = int(sys.argv[2])

wb = openpyxl.load_workbook(script)
sheet = wb.active
wb2 = openpyxl.Workbook()
sheet2 = wb2.active

for i in range(1, sheet.max_row+1):
    if i < m:
        for j in range(1, sheet.max_column+1):
            content = sheet.cell(row=i, column=j).value
            sheet2.cell(row=i, column=j).value = content
    
    else:
        for j in range(1, sheet.max_column+1):
            content = sheet.cell(row=i, column=j).value
            sheet2.cell(row=i+n, column=j).value = content

wb2.save('11_13_2.xlsx')
