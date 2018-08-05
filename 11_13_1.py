import openpyxl,sys

wb = openpyxl.Workbook()
sheet = wb.active

n = int(sys.argv[1])
for i in range(2, n+2):
    sheet.cell(row = i, column = 1).value = int(i) - 1

    sheet.cell(row = 1, column = i).value = int(i) - 1


for i in range(2, n+2):
    for j in range(2, n+2):
        sheet.cell(row =i, column = j).value = (int(i)-1)*(int(j)-1)

wb.save('11_13_1.xlsx')   
